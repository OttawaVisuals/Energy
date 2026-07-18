"""
add_oil_propane_rates.py

Adds heating-oil, propane, and heating-wood prices to
utility_rates_reference.json/.csv (built by utility_rates_reference.py), and
fills every remaining electricity/gas/oil/propane gap with a documented
fallback, so the table has one number per fuel for every province.

Heating oil / propane come from two NRCan "Prices by City" Excel exports:
  https://www2.nrcan.gc.ca/eneene/sources/pripri/prices_bycity_e.cfm
  (productID=7 "Furnace Oil", productID=6 "Auto Propane")

Unlike the electricity/gas source (a scraped JSON API), NRCan's by-city page
has no stable machine-readable endpoint — it's a manually-downloaded Excel
export per fuel. This script is therefore a manual-refresh step: re-download
both files from the URL above (same locationID query string selects "all
cities") and re-run whenever the reference table needs updating.

Both workbooks share one layout: a few title rows, a city-name header row
(each city spanning several columns — Price/Taxes[/Marketing/Refining
Margin]), a sub-header row, then one data row per period (monthly for
furnace oil, weekly for propane) with the most recent period last. Only the
"Price" column per city is used; the latest non-blank value per city is
taken (coverage varies — a city can have a blank latest week/month with
data in an earlier one), then averaged (simple mean, unweighted) across the
cities NRCan surveys in each province.

Cities NRCan surveys with no clear province mapping are flagged, not
guessed silently — see CITY_PROVINCE_NOTES.

Gap-filling (see FALLBACK_NOTES for the exact reasoning stored in the
output): a handful of provinces have no upstream data for a given fuel —
either because NRCan doesn't survey any city there (e.g. no furnace-oil
cities in AB/MB/SK/YT) or because the canada-utility-rates tariff scrape
has no residential gas utility for that province (e.g. PE, which has no
piped natural gas at all). Rather than leave those cells blank, each is
filled with a documented substitute and tagged low-confidence:
  - oil / propane: NRCan's own "Canada" national-average column from the
    same workbook (excluded from the per-city province averages above).
  - natural gas: the unweighted mean of this table's OTHER provinces'
    natural-gas rates (rates.json has no separate "Canada" figure to use).

Heating wood has no province-level source at all — see WOOD_CAD_PER_KWH's
derivation below — so the same flat $/kWh screening estimate is applied to
every province.

Usage:
    python add_oil_propane_rates.py <furnace_oil.xlsx> <auto_propane.xlsx>
"""

import datetime
import json
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_JSON = REPO_ROOT / "utility_rates_reference.json"
OUTPUT_CSV = REPO_ROOT / "utility_rates_reference.csv"

# NRCan's by-city page groups cities alphabetically with no province column;
# this maps every city in the current 72-city export to its province. Built
# from NRCan's known regional monitored-city sets (e.g. NB's 9 cities, NL's
# 5, Ontario's ~19) — see CITY_PROVINCE_NOTES for the one genuine ambiguity.
CITY_PROVINCE = {
    "Abbotsford": "BC", "Fort St. John": "BC", "Kamloops": "BC", "Kelowna": "BC",
    "Prince George": "BC", "Vancouver": "BC", "Victoria": "BC",
    "Calgary": "AB", "Edmonton": "AB", "Grande Prairie": "AB", "Lethbridge": "AB",
    "Lloydminster": "AB", "Red Deer": "AB",
    "Moose Jaw": "SK", "Prince Albert": "SK", "Regina": "SK", "Saskatoon": "SK",
    "Brandon": "MB", "Winnipeg": "MB",
    "Barrie": "ON", "Brantford": "ON", "Guelph": "ON", "Hamilton": "ON",
    "Kingston": "ON", "Kitchener": "ON", "London": "ON", "North Bay": "ON",
    "Oshawa": "ON", "Ottawa": "ON", "Peterborough": "ON", "Sarnia": "ON",
    "Sault Ste Marie": "ON", "St. Catharines": "ON", "Sudbury": "ON",
    "Thunder Bay": "ON", "Timmins": "ON", "Toronto": "ON", "Windsor": "ON",
    "Chicoutimi": "QC", "Drummondville": "QC", "Gaspé": "QC", "Gatineau": "QC",
    "Montreal": "QC", "Quebec": "QC", "Rimouski": "QC", "Sherbrooke": "QC",
    "Trois-Rivières": "QC", "Val d'Or": "QC",
    "Bathurst": "NB", "Campbellton": "NB", "Edmundston": "NB", "Fredericton": "NB",
    "Miramichi": "NB", "Moncton": "NB", "Saint John": "NB", "Sussex": "NB",
    "Woodstock": "NB",
    "Halifax": "NS", "Kentville": "NS", "New Glasgow": "NS", "Sydney": "NS",
    "Truro": "NS", "Yarmouth": "NS",
    "Charlottetown": "PE",
    "Corner Brook": "NL", "Gander": "NL", "Grand Falls": "NL",
    "Labrador City": "NL", "St. John's": "NL",
    "Whitehorse": "YT",
    "Yellowknife": "NT",
}
CITY_PROVINCE_NOTES = {
    "Lloydminster": "straddles the AB/SK border; assigned to AB here (unconfirmed against "
                    "NRCan's own convention) — treat this city's contribution as low-confidence.",
    "Grand Falls": "assumed to be Grand Falls-Windsor, NL (NL's standard 5-city monitored set), "
                   "not Grand Falls, NB.",
}

# --- Heating wood (pellets) ---------------------------------------------------
# No province-level pricing source was found; a single national figure from
# Canadian Biomass Magazine's "North America heating wood pellet market
# update, 2Q 2025" (https://www.canadianbiomassmagazine.ca/
# north-america-heating-wood-pellet-market-update-2q-2025/) is applied flat
# across every province. The article's own number ($290/ton) is embedded in
# a chart image, not machine-readable text, so the currency and ton
# definition below are ASSUMPTIONS, not confirmed from the source — flagged
# low-confidence accordingly.
WOOD_PELLET_CAD_PER_TON = 290.0
WOOD_PELLET_ASSUMPTIONS = (
    "Price is assumed CAD, assumed short ton (2,000 lb / 907.185 kg) — the unit chart values "
    "in the source article render as an image, so these weren't independently confirmed from "
    "article text.",
    "Energy content assumed at 16.5 million BTU/ton HHV, a standard figure for premium wood "
    "pellets (~8,250 BTU/lb) cited by pellet-industry fact sheets — this article doesn't state "
    "a heat content.",
    "Converted: 16,500,000 BTU/ton x 0.00029307107 kWh/BTU = 4,835.7 kWh/ton delivered heat "
    "content; $290 / 4,835.7 kWh = $0.0600/kWh.",
    "The source price may reflect wholesale/bulk market pricing rather than a residential "
    "retail rate; applying it to household heating cost is itself an approximation.",
)
BTU_PER_KWH = 3412.14
KWH_PER_WOOD_TON = 16_500_000 / BTU_PER_KWH
WOOD_CAD_PER_KWH = round(WOOD_PELLET_CAD_PER_TON / KWH_PER_WOOD_TON, 4)

FALLBACK_NOTES = {
    "oil_propane_canada_avg": "No NRCan city-level data for this province; filled with NRCan's "
                              "own 'Canada' national-average column from the same workbook.",
    "gas_canada_avg": "No residential gas utility in the canada-utility-rates tariff export for "
                      "this province (may reflect no piped natural gas service, e.g. PE); filled "
                      "with the unweighted mean of this table's other provinces' gas rates.",
}


def latest_price_per_city(ws, header_row_idx, cols_per_city, first_price_col=1):
    """{city: (value_cents_per_l, period)} using the last non-blank 'Price'
    reading per city. header_row_idx/period rows are 0-based into
    ws.iter_rows(values_only=True))."""
    rows = list(ws.iter_rows(values_only=True))
    header = rows[header_row_idx]
    # city name repeats at the start of its column block; walk forward to
    # find each block's starting column.
    city_cols = []
    col = first_price_col
    while col < len(header):
        name = header[col]
        if name:
            city_cols.append((name, col))
        col += cols_per_city
    data_rows = [r for r in rows[header_row_idx + 2:] if isinstance(r[0], datetime.datetime)]
    out = {}
    for city, col in city_cols:
        for r in reversed(data_rows):
            v = r[col]
            if isinstance(v, (int, float)):
                out[city] = (v, r[0].strftime("%Y-%m-%d"))
                break
    return out


def by_province(city_prices, unmapped):
    """Returns (per-province city lists, national 'Canada' (cents, period))."""
    prov = {}
    canada = city_prices.get("Canada")
    for city, (cents, period) in city_prices.items():
        if city == "Canada":
            continue
        p = CITY_PROVINCE.get(city)
        if p is None:
            unmapped.add(city)
            continue
        prov.setdefault(p, []).append((city, cents, period))
    return prov, canada


def build_fuel_entry(prov_cities):
    cents_vals = [c for _, c, _ in prov_cities]
    dollars_per_l = round((sum(cents_vals) / len(cents_vals)) / 100.0, 4)
    return {
        "cad_per_litre": dollars_per_l,
        "n_cities": len(prov_cities),
        "source": "province_cities",
        "sources": [
            {"city": city, "cents_per_litre": cents, "period": period}
            for city, cents, period in sorted(prov_cities)
        ],
    }


def build_fuel_entry_canada_avg(cents, period):
    return {
        "cad_per_litre": round(cents / 100.0, 4),
        "n_cities": 0,
        "source": "canada_average_nrcan",
        "notes": [FALLBACK_NOTES["oil_propane_canada_avg"]],
        "sources": [{"city": "Canada (national average)", "cents_per_litre": cents, "period": period}],
    }


def main():
    if len(sys.argv) != 3:
        print("usage: python add_oil_propane_rates.py <furnace_oil.xlsx> <auto_propane.xlsx>",
              file=sys.stderr)
        sys.exit(1)
    oil_path, propane_path = sys.argv[1], sys.argv[2]

    if not OUTPUT_JSON.exists():
        raise RuntimeError(f"{OUTPUT_JSON} not found — run utility_rates_reference.py first")
    doc = json.loads(OUTPUT_JSON.read_text(encoding="utf-8"))

    unmapped = set()

    wb = openpyxl.load_workbook(oil_path, data_only=True)
    oil_prices = latest_price_per_city(wb["Sheet1"], header_row_idx=3, cols_per_city=4)
    oil_by_prov, oil_canada = by_province(oil_prices, unmapped)

    wb = openpyxl.load_workbook(propane_path, data_only=True)
    propane_prices = latest_price_per_city(wb["Sheet1"], header_row_idx=3, cols_per_city=2)
    propane_by_prov, propane_canada = by_province(propane_prices, unmapped)

    if unmapped:
        raise RuntimeError(f"unmapped cities (add to CITY_PROVINCE): {sorted(unmapped)}")

    all_provinces = sorted(doc["provinces"])

    # --- oil / propane: province cities, else NRCan's own Canada average ---
    for p in all_provinces:
        entry = doc["provinces"][p]
        entry["heating_oil"] = (build_fuel_entry(oil_by_prov[p]) if p in oil_by_prov
                                else build_fuel_entry_canada_avg(*oil_canada))
        entry["propane"] = (build_fuel_entry(propane_by_prov[p]) if p in propane_by_prov
                            else build_fuel_entry_canada_avg(*propane_canada))

    # --- natural gas: fill provinces with no upstream tariff using the mean
    # of the other provinces already in the table ---
    gas_provinces_real = [p for p in all_provinces if "natural_gas" in doc["provinces"][p]]
    gas_vals = [doc["provinces"][p]["natural_gas"]["dollars_per_m3"] for p in gas_provinces_real]
    gas_canada_avg = round(sum(gas_vals) / len(gas_vals), 4)
    for p in all_provinces:
        entry = doc["provinces"][p]
        if "natural_gas" not in entry:
            entry["natural_gas"] = {
                "dollars_per_m3": gas_canada_avg,
                "n_utilities": 0,
                "confidence": "low",
                "source": "canada_average_table",
                "notes": [FALLBACK_NOTES["gas_canada_avg"]],
                "sources": [],
            }
        else:
            entry["natural_gas"]["source"] = "province_utilities"

    # --- heating wood: flat national estimate, every province ---
    for p in all_provinces:
        doc["provinces"][p]["heating_wood"] = {
            "cad_per_kwh": WOOD_CAD_PER_KWH,
            "source": "canadian_biomass_magazine_flat",
            "confidence": "low",
            "notes": list(WOOD_PELLET_ASSUMPTIONS),
        }

    if "url" in doc.get("source", {}):  # first run after this script: flatten -> {tariffs: {...}}
        doc["source"] = {"tariffs": doc["source"]}
    doc["source"]["heating_oil_propane"] = {
        "name": "NRCan Prices by City (Furnace Oil / Auto Propane)",
        "url": "https://www2.nrcan.gc.ca/eneene/sources/pripri/prices_bycity_e.cfm",
        "cadence": "manual re-download (no stable API); furnace oil monthly, propane weekly",
        "method": "simple mean across NRCan-surveyed cities per province, latest available "
                  "reading per city (coverage varies by city); provinces with no surveyed "
                  "cities use NRCan's own 'Canada' national-average column instead",
        "city_province_caveats": CITY_PROVINCE_NOTES,
    }
    doc["source"]["heating_wood"] = {
        "name": "Canadian Biomass Magazine — North America heating wood pellet market update, 2Q 2025",
        "url": "https://www.canadianbiomassmagazine.ca/north-america-heating-wood-pellet-market-update-2q-2025/",
        "cadence": "one-off (not re-fetched automatically)",
        "method": f"single flat ${WOOD_PELLET_CAD_PER_TON:.0f}/ton figure applied to every "
                  "province — no province-level wood price source was found",
        "assumptions": list(WOOD_PELLET_ASSUMPTIONS),
    }
    doc["source"]["natural_gas_fallback"] = {
        "method": FALLBACK_NOTES["gas_canada_avg"],
        "canada_average_dollars_per_m3": gas_canada_avg,
        "provinces_averaged": sorted(gas_provinces_real),
    }
    doc["units"]["heating_oil"] = "cad_per_litre"
    doc["units"]["propane"] = "cad_per_litre"
    doc["units"]["heating_wood"] = "cad_per_kwh (delivered heat content)"

    OUTPUT_JSON.write_text(json.dumps(doc, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"wrote {OUTPUT_JSON}")

    provinces = all_provinces
    csv_lines = ["province,elec_cents_per_kwh,elec_fixed_monthly_cad,elec_n_utilities,elec_confidence,"
                 "gas_dollars_per_m3,gas_fixed_monthly_cad,gas_source,"
                 "heating_oil_cad_per_litre,heating_oil_source,"
                 "propane_cad_per_litre,propane_source,"
                 "heating_wood_cad_per_kwh"]
    for p in provinces:
        entry = doc["provinces"][p]
        e, g = entry.get("electricity"), entry.get("natural_gas")
        o, pr, w = entry.get("heating_oil"), entry.get("propane"), entry.get("heating_wood")
        csv_lines.append(",".join([
            p,
            f"{e['cents_per_kwh']:.3f}" if e else "",
            f"{e['fixed_monthly_cad']:.2f}" if e else "",
            str(e["n_utilities"]) if e else "",
            e["confidence"] if e else "",
            f"{g['dollars_per_m3']:.4f}",
            f"{g.get('fixed_monthly_cad', ''):.2f}" if g.get("fixed_monthly_cad") is not None else "",
            g["source"] if "source" in g else "province_utilities",
            f"{o['cad_per_litre']:.4f}",
            o["source"],
            f"{pr['cad_per_litre']:.4f}",
            pr["source"],
            f"{w['cad_per_kwh']:.4f}",
        ]))
    OUTPUT_CSV.write_text("\n".join(csv_lines) + "\n", encoding="utf-8")
    print(f"wrote {OUTPUT_CSV}")

    print("\noil / propane / wood by province:")
    for p in provinces:
        entry = doc["provinces"][p]
        o, pr, w = entry["heating_oil"], entry["propane"], entry["heating_wood"]
        print(f"  {p}  oil=${o['cad_per_litre']:.4f}/L ({o['source']})  "
              f"propane=${pr['cad_per_litre']:.4f}/L ({pr['source']})  "
              f"wood=${w['cad_per_kwh']:.4f}/kWh")
    print(f"\ngas Canada-average fallback: ${gas_canada_avg:.4f}/m3 "
          f"(used for {sum(1 for p in provinces if doc['provinces'][p]['natural_gas']['source']=='canada_average_table')} provinces)")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\nERROR: {e}", file=sys.stderr)
        sys.exit(1)
