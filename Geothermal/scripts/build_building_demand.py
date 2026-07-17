"""
build_building_demand.py  --  Heat Demand Phase 2 (HEATDEMAND_PLAN.md §2-4)

Adds, to every building in Data/processed/buildings_ottawa.parquet, a
*screening estimate* of:
    annual_kwh   -- annual SPACE-HEAT delivered energy (kWh/yr)
    design_kw    -- design-day heat loss (kW at Ottawa 2.5%-ile design temp)
    ua_w_per_k   -- effective envelope UA (W/K) -- feeds Phase 3 electrification
    heat_fuel    -- primary heating fuel (gas/electric/oil/propane/wood)
    plus method/confidence flags and an EWRB-actuals cross-check column.

This is a SCREENING LAYER, not a set of building audits: vintage and the
ambiguous-type class split are probabilistic per building (Phase 1), intensities
are population medians, and the fuel is a raked probabilistic draw. Numbers are
meaningful in aggregate (500 m cell / feeder, Phase 4), not for a single address.

METHOD SUMMARY (full write-up: Geothermal/README.md §3.11)
  Residential houses (detached, row)
      Ottawa ERS archetypes (HeatPump/data/processed/archetypes.json, vintage x
      type), annual kWh + design kW + UA scaled by floor-area ratio
      clip(floor_area / archetype_floor_area, 0.5, 2.5). 'row' folds in semi/
      duplex (Phase 1's class scheme has no 'semi' bucket) and uses the
      townhouse_row archetype -- validated against CEUD ON single_attached
      (row+semi+attached) which agrees to a few %.
  Apartments / MURBs (lowrise_murb, highrise_murb)
      Ottawa apartment space-heat intensity per m2 built by transferring the
      CEUD Ontario apartment/single-attached per-m2 RATIO onto the Ottawa
      townhouse_row archetype per-m2 intensity -- keeps apartments internally
      consistent with (more efficient than) the Ottawa row house and needs no
      external HDD guess. floor_area = footprint x storeys.
  Non-residential (commercial, institutional, industrial)
      EWRB-2024 Ottawa ACTUAL median Site_EUI by property-type group (Ottawa
      actuals, not CEUD's provincial commercial intensity which is ~1.9x higher
      -- a known CEUD commercial floor-space undercount, confirmed here) x the
      CEUD Ontario commercial SPACE-HEAT share (~0.57) to isolate heating.
      Industrial is a clearly-flagged low-confidence placeholder (mostly
      agricultural/small-industrial footprints, process-not-space-heat).
  Fuel
      ERS per-FSA pre-retrofit Pre_HeatFuel mix, raked by iterative proportional
      fitting to the StatCan 38-10-0286 Ottawa-ON-part shares (gas 59 %,
      electric 21 %; oil/propane/wood tail is suppressed at that geography so its
      internal split comes from ERS), with a rural no-natural-gas constraint from
      the City serviced-area layer, then a seeded per-building categorical draw.

Run from C:\\Energy:  python Geothermal/scripts/build_building_demand.py
Reads:  Data/processed/buildings_ottawa.parquet (Phase 1 output; augmented in place)
Writes: same .parquet (+ .gpkg) with the new columns and parquet metadata.
"""

import json
from pathlib import Path

import numpy as np
import pandas as pd
import geopandas as gpd

# ---------------------------------------------------------------------------
# Paths (project convention: run from C:\Energy)
# ---------------------------------------------------------------------------
ROOT       = Path(".").resolve()
BUILDINGS  = ROOT / "Geothermal/Data/processed/buildings_ottawa.parquet"
BUILD_GPKG = ROOT / "Geothermal/Data/processed/buildings_ottawa.gpkg"
ARCHETYPES = ROOT / "HeatPump/data/processed/archetypes.json"
TMY        = ROOT / "HeatPump/data/processed/tmy_temps.json"
CEUD_RES   = ROOT / "ceud_json/res_on.json"
CEUD_COM   = ROOT / "ceud_json/com_on.json"
ERS_ON     = Path(r"C:/ERS/web/ers_web_ON.parquet")
SERVICED   = ROOT / "Geothermal/Data/processed/city_open_loop_potential.geojson"
EWRB       = ROOT / "Geothermal/Data/Raw/EWRB_2024.xlsx"
TANDM      = ROOT / "Geothermal/Data/Raw/references/TaNDM_Kelowna_2021_AnalysisReady.xlsx"
DA_CENSUS  = ROOT / "Geothermal/Data/processed/da_census.json"

SEED = 20260716            # same fixed seed convention as Phase 1
KWH_PER_GJ = 1000.0 / 3.6  # 277.778

# footprint x storeys is GROSS external floor area (includes attached garage,
# external wall thickness, and a roof-pitch storey over-count where height/3
# rounds a steep-roofed house up a storey). The ERS archetype floor_area is
# HEATED interior area, so the house method converts gross->heated before taking
# the archetype scale. 0.80 = ~15% garage/wall + ~5% roof-pitch over-count;
# screening assumption, documented in README §3.11. Chosen so the modelled
# per-UNIT detached mean matches CEUD ON single_detached (the like-for-like
# check), NOT to force the city-wide total -- that total runs high for a
# separate, documented reason (Phase 1 over-attributes detached/MURB
# buildings; see validation (a) and README §3.10).
HEATED_FRACTION_HOUSE = 0.80
CEUD_YEAR  = 2021          # census-vintage year for the stock; CEUD has 2000-2023

# Ottawa heating design conditions (HeatPump/METHODOLOGY.md Phase 4).
T_SET_C    = 21.0          # HOT2000 indoor design setpoint
T_DESIGN_C = -22.8         # Ottawa 2.5%-ile January (METHODOLOGY.md)
DELTA_T_DESIGN = T_SET_C - T_DESIGN_C   # 43.8 K

# StatCan 38-10-0286, Ottawa-Gatineau (Ontario part, member 24), 2023 -- the
# latest published year. Recorded in Geothermal/Data/heatdemand_source_notes.md
# §2.2 (the full-table CSV is not cached; these two shares are the only ones
# StatCan publishes at this geography -- oil/wood/propane/heat-pump are
# suppressed, so the tail's internal split is taken from ERS).
STATCAN_GAS_SHARE  = 0.59
STATCAN_ELEC_SHARE = 0.21

FUELS = ["gas", "electric", "oil", "propane", "wood"]

# Equivalent full-load heating hours by balance point, computed from the Ottawa
# TMY at load-time (EFLH = HDH(Tbal) / DELTA_T_DESIGN). Houses use their
# archetype's own annual/design ratio; the categories below pick a balance point
# (see README §3.11). Filled in main() from the TMY so they stay reproducible.
BALANCE_POINT_C = {          # assumed heating balance point per non-house class
    "apartment":     10.0,   # ~ Ottawa detached archetype (moderate gains)
    "commercial":    12.0,   # continuous heating + higher internal gains
    "institutional": 12.0,
    "industrial":     8.0,   # lightly heated, low off-hours gains
}

# Heating-system seasonal efficiency used to turn delivered space heat into
# fuel input for the gas-share / emissions bookkeeping (screening).
AFUE = {"gas": 0.92, "oil": 0.83, "propane": 0.85, "wood": 0.65, "electric": 1.0}

# Emission factors (kg CO2e per kWh of *fuel input*) for the community-inventory
# cross-check only. NG: 1.921 kg/m3 / 10.33 kWh/m3 = 0.186; ON grid ~ 0.030
# (hydro/nuclear dominated); oil 0.267; propane 0.214; wood ~ biogenic (0).
EF_KG_PER_KWH = {"gas": 0.186, "electric": 0.030, "oil": 0.267,
                 "propane": 0.214, "wood": 0.0}


# ===========================================================================
# Loaders / small helpers
# ===========================================================================
def load_archetypes():
    a = json.load(open(ARCHETYPES))["Ottawa"]
    # per-archetype floor-area-normalised space-heat intensity (kWh/m2) is used
    # to build the apartment intensity below.
    return a


def vintage_to_arch(vint: str) -> str:
    """Phase 1's 8 census bands -> the 3 archetype detached vintage bins."""
    if vint in ("1960_or_before", "1961_1980"):
        return "pre_1980_detached"
    if vint in ("1981_1990", "1991_2000", "2001_2005"):
        return "1980_2005_detached"
    return "post_2005_detached"   # 2006_2010, 2011_2015, 2016_2021


def ceud_record(recs, year, building_type, end_use):
    # segment2 is the commercial activity dimension; require it None so the
    # commercial *grand total* isn't confused with a single activity's total.
    for r in recs:
        if (r["year"] == year and r.get("building_type") == building_type
                and r.get("end_use") == end_use
                and r.get("energy_source") is None
                and r.get("segment2") is None):
            return r["energy_PJ"]
    return None


def ceud_expl(expl, year, variable, segment):
    for r in expl:
        if r["year"] == year and r["variable"] == variable and r["segment"] == segment:
            return r["value"]
    return None


def build_intensities(arch):
    """Return a dict of per-m2 space-heat intensities (kWh/m2/yr) and the
    supporting CEUD/EWRB numbers, all derived live so the script is reproducible.
    """
    res = json.load(open(CEUD_RES))
    com = json.load(open(CEUD_COM))

    # --- apartments: CEUD ON apartment/single_attached per-m2 ratio x Ottawa row
    apt_sh   = ceud_record(res["records"], CEUD_YEAR, "apartments", "space_heating")
    att_sh   = ceud_record(res["records"], CEUD_YEAR, "single_attached", "space_heating")
    apt_fs   = ceud_expl(res["explanatory"], CEUD_YEAR, "floor_space", "apartments")
    att_fs   = ceud_expl(res["explanatory"], CEUD_YEAR, "floor_space", "single_attached")
    apt_hh   = ceud_expl(res["explanatory"], CEUD_YEAR, "households", "apartments")
    # per-m2 (GJ/m2): PJ*1e6 GJ / (Mm2 * 1e6 m2)
    apt_gj_m2 = apt_sh / apt_fs
    att_gj_m2 = att_sh / att_fs
    ceud_apt_att_ratio = apt_gj_m2 / att_gj_m2
    row = arch["townhouse_row"]
    row_kwh_m2 = row["annual_heat_kWh"] / row["floor_area_m2"]
    apt_kwh_m2 = row_kwh_m2 * ceud_apt_att_ratio
    gross_m2_per_unit = (apt_fs * 1e6) / (apt_hh * 1e3)   # for the units estimate

    # --- commercial space-heat share (CEUD ON, aggregate end-use) ------------
    com_sh  = ceud_record(com["records"], CEUD_YEAR, None, "space_heating")
    com_tot = ceud_record(com["records"], CEUD_YEAR, None, None)
    com_fs  = ceud_expl(com["explanatory"], CEUD_YEAR, "total_floor_space", "all")
    com_sh_share = com_sh / com_tot
    ceud_com_gj_m2 = com_sh / com_fs                      # CEUD's own (inflated)

    # --- EWRB Ottawa actual total Site_EUI (GJ/m2) by class group ------------
    ewrb = ewrb_medians()

    # non-residential space-heat intensity = EWRB actual total EUI x space share
    com_kwh_m2  = ewrb["commercial"]    * com_sh_share * KWH_PER_GJ
    inst_kwh_m2 = ewrb["institutional"] * com_sh_share * KWH_PER_GJ
    # industrial: EWRB warehouse actuals are large heated DCs; our 'industrial'
    # class is mostly agricultural/small-industrial with minimal space heat, so
    # apply a lower space-heat share (0.40) and flag low confidence.
    ind_kwh_m2  = ewrb["industrial"]    * 0.40           * KWH_PER_GJ

    return {
        "apt_kwh_m2": apt_kwh_m2,
        "com_kwh_m2": com_kwh_m2,
        "inst_kwh_m2": inst_kwh_m2,
        "ind_kwh_m2": ind_kwh_m2,
        "gross_m2_per_unit": gross_m2_per_unit,
        "_diag": {
            "ceud_apt_att_ratio": ceud_apt_att_ratio,
            "row_kwh_m2": row_kwh_m2,
            "ceud_com_space_share": com_sh_share,
            "ceud_com_gj_m2": ceud_com_gj_m2,
            "ewrb": ewrb,
        },
    }


def ewrb_medians():
    """Median Ottawa Site_EUI (GJ/m2, total all-end-use) by class group from the
    EWRB-2024 workbook. EWRB has no street address (source notes §2.1) so it can
    only calibrate intensities, not override individual buildings."""
    import openpyxl
    wb = openpyxl.load_workbook(EWRB, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(h) for h in next(rows)]
    idx = {h: i for i, h in enumerate(hdr)}
    ci, pt, se = idx["City"], idx["PrimPropTypCalc"], idx["Site_EUI1"]
    groups = {
        "commercial": {"Office", "Retail Store", "Strip Mall", "Mixed Use Property",
                       "Enclosed Mall", "Other - Mall", "Supermarket/Grocery Store",
                       "Financial Office", "Hotel", "Wholesale Club/Supercenter",
                       "Vehicle Dealership", "Mailing Center/Post Office", "Parking",
                       "Other", "Other - Services"},
        "institutional": {"Senior Living Community", "K-12 School", "College/University",
                          "Worship Facility", "Library", "Medical Office",
                          "Residence Hall/Dormitory", "Residential Care Facility",
                          "Other - Specialty Hospital", "Hospital (General Medical & Surgical)",
                          "Ambulatory Surgical Center", "Other - Lodging/Residential"},
        "industrial": {"Non-Refrigerated Warehouse", "Distribution Center",
                       "Refrigerated Warehouse", "Self-Storage Facility",
                       "Manufacturing/Industrial Plant"},
    }
    buckets = {k: [] for k in groups}
    for r in rows:
        if not r[ci] or "ottawa" not in str(r[ci]).lower():
            continue
        v = r[se]
        if not isinstance(v, (int, float)):
            continue
        for gname, types in groups.items():
            if str(r[pt]) in types:
                buckets[gname].append(float(v))
                break
    # fallbacks if a bucket is thin
    med = {}
    for k, v in buckets.items():
        med[k] = float(np.median(v)) if v else float("nan")
        med[k + "_n"] = len(v)
    if med["institutional_n"] < 8:      # small sample -> lean on commercial
        med["institutional"] = med["commercial"] * 0.9
    return med


# ===========================================================================
# Fuel assignment
# ===========================================================================
def ers_fsa_fuel_shares(min_rows=30):
    """Per-FSA pre-retrofit fuel probability vectors from ERS ON (K-prefix FSAs),
    plus an Ottawa-wide fallback for thin FSAs."""
    df = pd.read_parquet(ERS_ON, columns=["FSA", "Pre_HeatFuel"]).dropna()
    df["FSA"] = df["FSA"].astype(str).str.upper()
    df = df[df["FSA"].str.match(r"^K[0-9]")]
    fmap = {"Natural Gas": "gas", "Electricity": "electric", "Oil": "oil",
            "Propane": "propane", "Mixed Wood": "wood", "Hardwood": "wood",
            "Wood Pellets": "wood"}
    df["fuel"] = df["Pre_HeatFuel"].map(fmap)
    df = df.dropna(subset=["fuel"])

    def shares(sub):
        c = sub["fuel"].value_counts()
        return np.array([c.get(f, 0) for f in FUELS], float)

    city = shares(df)
    city = city / city.sum()
    per_fsa = {}
    for fsa, sub in df.groupby("FSA"):
        if len(sub) >= min_rows:
            v = shares(sub)
            per_fsa[fsa] = v / v.sum()
    return per_fsa, city


def assign_fuel(bldg, serviced, per_fsa, city_share, rng, city_mask):
    """Rake ERS per-FSA fuel probabilities to the StatCan gas/electric targets
    (residential classes only) under a rural no-gas constraint via IPF, then draw.
    Non-residential classes use a documented serviced/rural prior (StatCan
    38-10-0286 is a *dwelling* table, so it does not constrain non-res).

    Phase 2.5: the StatCan 38-10-0286 target (gas 59 % / electric 21 %) is for
    the Ottawa-Gatineau Ontario-part CMA -- i.e. the CITY, not the bbox. So the
    IPF is fitted over the in_ottawa_cd residential subset (city_mask). Rural
    residential buildings OUTSIDE the CD keep their per-FSA / no-gas mix and are
    not part of the city target (they were never in the CMA the target
    describes); including them in the rake used to hit 59/21 bbox-wide while
    leaving the city subset too gassy."""
    n = len(bldg)
    P = np.zeros((n, len(FUELS)))
    fsa_arr = bldg["fsa"].fillna("").astype(str).str.upper().values
    for i, f in enumerate(fsa_arr):
        P[i] = per_fsa.get(f, city_share)
    # rural (unserviced) buildings cannot have natural gas
    gi = FUELS.index("gas")
    rural = ~serviced
    P[rural, gi] = 0.0
    # a rural building whose FSA ERS mix is ~100% gas now has an all-zero row --
    # fall back to the city-wide NON-gas distribution rather than leaving NaN
    # (a NaN row would otherwise be silently drawn as fuel index 0 = gas).
    nogas_city = city_share.copy()
    nogas_city[gi] = 0.0
    nogas_city = nogas_city / nogas_city.sum()
    zero_rows = P.sum(axis=1) < 1e-12
    P[zero_rows] = nogas_city
    P = P / P.sum(axis=1, keepdims=True)

    res_mask = bldg["class"].isin(["detached", "row", "lowrise_murb", "highrise_murb"]).values
    rake_mask = res_mask & np.asarray(city_mask)   # city residential = the CMA the target describes

    # --- IPF over the CITY residential subset --------------------------------
    ei = FUELS.index("electric")
    tail = ["oil", "propane", "wood"]
    # tail internal split from the city ERS mix, scaled to (1 - gas - elec)
    tail_base = np.array([city_share[FUELS.index(t)] for t in tail])
    tail_base = tail_base / tail_base.sum()
    tail_total = 1.0 - STATCAN_GAS_SHARE - STATCAN_ELEC_SHARE
    target = np.zeros(len(FUELS))
    target[gi] = STATCAN_GAS_SHARE
    target[ei] = STATCAN_ELEC_SHARE
    for t, s in zip(tail, tail_base):
        target[FUELS.index(t)] = tail_total * s

    Pr = P[rake_mask].copy()
    m = np.ones(len(FUELS))
    for _ in range(200):
        Q = Pr * m
        Q = Q / Q.sum(axis=1, keepdims=True)
        cur = Q.mean(axis=0)
        # multiplicative update; skip fuels with no achievable mass
        ratio = np.where(cur > 1e-9, target / cur, 1.0)
        m = m * ratio
        if np.max(np.abs(cur - target)) < 1e-4:
            break
    Q = Pr * m
    Q = Q / Q.sum(axis=1, keepdims=True)
    P[rake_mask] = Q
    # outside-CD residential keep their per-FSA / no-gas P (not re-weighted)
    print("  IPF city-residential expected shares: " +
          ", ".join(f"{f} {Q.mean(axis=0)[i]*100:.1f}%" for i, f in enumerate(FUELS))
          + f"  (rural-forced-nongas share of city res: "
          f"{(rural & rake_mask).sum()/rake_mask.sum()*100:.1f}%)")

    # --- non-residential prior -----------------------------------------------
    nonres = ~res_mask
    if nonres.any():
        serv = serviced & nonres
        rur = (~serviced) & nonres
        prior_serv = np.array([0.85, 0.15, 0.0, 0.0, 0.0])            # gas/elec
        prior_rur = np.array([0.0, 0.10, 0.45, 0.45, 0.0])            # oil/propane/elec
        P[serv] = prior_serv
        P[rur] = prior_rur

    # --- seeded per-building categorical draw --------------------------------
    P = P / P.sum(axis=1, keepdims=True)
    u = rng.random(n)
    cdf = np.cumsum(P, axis=1)
    draw = (u[:, None] < cdf).argmax(axis=1)
    return np.array(FUELS)[draw]


# ===========================================================================
# Main
# ===========================================================================
def main():
    print("=" * 74)
    print("Heat Demand Phase 2 -- build_building_demand.py")
    print("=" * 74)
    rng = np.random.default_rng(SEED)

    arch = load_archetypes()
    intens = build_intensities(arch)
    d = intens["_diag"]
    print("\nDerived space-heat intensities (kWh/m2/yr):")
    print(f"  apartment (MURB)   {intens['apt_kwh_m2']:6.1f}   "
          f"(= Ottawa row {d['row_kwh_m2']:.1f} x CEUD apt/attached ratio "
          f"{d['ceud_apt_att_ratio']:.3f})")
    print(f"  commercial         {intens['com_kwh_m2']:6.1f}   "
          f"(EWRB {d['ewrb']['commercial']:.3f} GJ/m2 x space share "
          f"{d['ceud_com_space_share']:.3f})")
    print(f"  institutional      {intens['inst_kwh_m2']:6.1f}   "
          f"(EWRB {d['ewrb']['institutional']:.3f} GJ/m2)")
    print(f"  industrial         {intens['ind_kwh_m2']:6.1f}   (low confidence)")
    print(f"  [CEUD raw commercial space-heat {d['ceud_com_gj_m2']*KWH_PER_GJ:.1f} "
          f"kWh/m2 -- {d['ceud_com_gj_m2']/ (d['ewrb']['commercial']*d['ceud_com_space_share']):.2f}x "
          f"the EWRB-calibrated value; CEUD floor-space undercount, not used]")

    # EFLH per balance point from the Ottawa TMY (reproducible) ---------------
    T = np.array(json.load(open(TMY))["Ottawa"], float)
    eflh = {k: float(np.sum(np.maximum(0.0, tb - T)) / DELTA_T_DESIGN)
            for k, tb in BALANCE_POINT_C.items()}
    print("\nEFLH from Ottawa TMY:", {k: round(v) for k, v in eflh.items()})

    # ---- load buildings ------------------------------------------------------
    print(f"\nLoading {BUILDINGS} ...")
    gdf = gpd.read_parquet(BUILDINGS)
    n = len(gdf)
    print(f"  {n:,} buildings")
    gdf["floor_area_m2"] = gdf["footprint_m2"] * gdf["storeys"]

    # ---- serviced-area flag (rural no-gas) ----------------------------------
    print("Flagging serviced (gas-available) buildings ...")
    serv = gpd.read_file(SERVICED).to_crs(4326)[["geometry"]]
    pts = gpd.GeoDataFrame(
        {"i": np.arange(n)},
        geometry=gdf.geometry.representative_point(), crs=4326)
    hit = gpd.sjoin(pts, serv, how="inner", predicate="within")["i"].unique()
    served = np.zeros(n, bool)
    served[hit] = True
    print(f"  serviced: {served.mean()*100:.1f}% of buildings")

    # ---- per-building space heat + design kW --------------------------------
    print("Modelling annual space heat + design kW ...")
    annual = np.zeros(n)
    design = np.zeros(n)
    ua = np.zeros(n)
    method = np.empty(n, dtype=object)
    conf = np.empty(n, dtype=object)
    units = np.zeros(n, dtype=int)

    cls = gdf["class"].values
    vint = gdf["vintage"].values
    fa = gdf["floor_area_m2"].values

    for i in range(n):
        c = cls[i]
        area = fa[i]
        if c in ("detached", "row"):
            akey = "townhouse_row" if c == "row" else vintage_to_arch(vint[i])
            a = arch[akey]
            heated = area * HEATED_FRACTION_HOUSE
            scale = min(2.5, max(0.5, heated / a["floor_area_m2"]))
            annual[i] = a["annual_heat_kWh"] * scale
            design[i] = a["design_heat_loss_kW"] * scale
            ua[i] = a["UA_W_per_K"] * scale
            method[i] = f"archetype:{akey}"
            conf[i] = "medium"
        elif c in ("lowrise_murb", "highrise_murb"):
            annual[i] = area * intens["apt_kwh_m2"]
            design[i] = annual[i] / eflh["apartment"]
            ua[i] = design[i] * 1000.0 / DELTA_T_DESIGN
            units[i] = max(1, round(area / intens["gross_m2_per_unit"]))
            method[i] = "ceud_apartment_intensity"
            conf[i] = "medium"
        elif c in ("commercial", "institutional"):
            im = intens["com_kwh_m2"] if c == "commercial" else intens["inst_kwh_m2"]
            annual[i] = area * im
            design[i] = annual[i] / eflh[c]
            ua[i] = design[i] * 1000.0 / DELTA_T_DESIGN
            method[i] = "ewrb_actual_x_ceud_share"
            conf[i] = "low"
        elif c == "accessory":
            # Phase 2.5 reconciliation class: probable non-dwelling structures
            # (garages/sheds/secondary buildings) reclassified because the DA's
            # census dwelling count cannot support them (build_building_stock.py
            # reconcile_stock, README §3.10). Treated as unheated for this
            # screening layer -- excluded from dwelling and residential tallies.
            annual[i] = 0.0
            design[i] = 0.0
            ua[i] = 0.0
            method[i] = "accessory_nonheated"
            conf[i] = "na"
        else:  # industrial
            annual[i] = area * intens["ind_kwh_m2"]
            design[i] = annual[i] / eflh["industrial"]
            ua[i] = design[i] * 1000.0 / DELTA_T_DESIGN
            method[i] = "industrial_placeholder"
            conf[i] = "very_low"

    gdf["annual_kwh"] = np.round(annual, 1)
    gdf["design_kw"] = np.round(design, 3)
    gdf["ua_w_per_k"] = np.round(ua, 1)
    gdf["units_est"] = units
    gdf["demand_method"] = method
    gdf["demand_confidence"] = conf

    # ---- fuel ---------------------------------------------------------------
    print("Assigning heating fuel (ERS per-FSA mix, raked to StatCan) ...")
    per_fsa, city_share = ers_fsa_fuel_shares()
    print(f"  city ERS fuel mix: " +
          ", ".join(f"{f} {city_share[i]*100:.0f}%" for i, f in enumerate(FUELS)))
    city_mask = (gdf["in_ottawa_cd"].to_numpy() if "in_ottawa_cd" in gdf.columns
                 else np.ones(len(gdf), bool))
    gdf["heat_fuel"] = assign_fuel(gdf, served, per_fsa, city_share, rng, city_mask)

    # ---- validation ---------------------------------------------------------
    validate(gdf, arch, intens, served)

    # ---- write --------------------------------------------------------------
    write_outputs(gdf)
    print("\nDone.")


# ===========================================================================
# Validation  (prints (a)-(e); investigate before shipping)
# ===========================================================================
def validate(gdf, arch, intens, served):
    print("\n" + "=" * 74)
    print("VALIDATION")
    print("=" * 74)
    # Phase 2.5: every CITY-WIDE sum is taken over the Ottawa census division
    # only (in_ottawa_cd) -- the bbox extends into surrounding townships but the
    # 2021 census household count is city-only, so bbox-wide sums are an
    # apples-to-oranges over-count (README §3.10). Fall back to all rows if the
    # column is absent (pre-Phase-2.5 parquet).
    if "in_ottawa_cd" in gdf.columns:
        city = gdf[gdf["in_ottawa_cd"]]
        print(f"City scope: {len(city):,}/{len(gdf):,} buildings inside the "
              f"Ottawa CD (in_ottawa_cd); city-wide sums use this subset.")
    else:
        city = gdf
        print("WARNING: no in_ottawa_cd column (pre-Phase-2.5 stock) -- city "
              "sums fall back to the full bbox and will over-count.")
    gdf_full = gdf
    gdf = city
    res = gdf[gdf["class"].isin(["detached", "row", "lowrise_murb", "highrise_murb"])]
    res_twh = res["annual_kwh"].sum() / 1e9
    total_twh = gdf["annual_kwh"].sum() / 1e9

    # --- (a) residential space heat vs CEUD ON scaled to Ottawa households ----
    r = json.load(open(CEUD_RES))
    on_sh_pj = ceud_record(r["records"], CEUD_YEAR, None, "space_heating")   # 315
    on_hh_k = ceud_expl(r["explanatory"], CEUD_YEAR, "households", "all")     # thousands
    on_gj_per_hh = on_sh_pj * 1e6 / (on_hh_k * 1e3)
    da = json.load(open(DA_CENSUS))
    ott_hh = sum((v.get("total_dwellings") or 0) for v in da.values())
    # Ottawa colder than ON average -> HDD uplift Ottawa/ON(populated).
    hdd_uplift = 4407.0 / 3900.0     # Ottawa TMY HDD18 4407; ON populated ~3900
    ceud_ott_twh = on_gj_per_hh * ott_hh * hdd_uplift * KWH_PER_GJ / 1e9  # GJ -> TWh
    print(f"\n(a) Residential space heat (Ottawa CD): modelled {res_twh:.2f} TWh "
          f"vs CEUD-scaled {ceud_ott_twh:.2f} TWh "
          f"(delta {100*(res_twh-ceud_ott_twh)/ceud_ott_twh:+.0f}%, target +-20%)  "
          f"[Ottawa dwellings {ott_hh:,}, CEUD ON {on_gj_per_hh:.1f} GJ/hh x "
          f"HDD uplift {hdd_uplift:.2f}]")
    # Phase 2.5: implied dwellings are now taken over the Ottawa CD (in_ottawa_cd)
    # and constrained per-DA to the census count (build_building_stock.py
    # reconcile_stock, README §3.10). The former 1.89x over-count was dominated
    # by the bbox extending beyond the city (~half the excess) plus MURB
    # no-height unit inflation; both are corrected in the stock, not by retuning
    # intensities -- the per-UNIT detached mean must stay ~= CEUD.
    murb = gdf[gdf["class"].isin(["lowrise_murb", "highrise_murb"])]
    implied = (gdf["class"] == "detached").sum() + (gdf["class"] == "row").sum() \
        + murb["units_est"].sum()
    det = gdf[gdf["class"] == "detached"]
    ceud_det_kwh = ceud_record(r["records"], CEUD_YEAR, "single_detached", "space_heating") \
        * 1e6 / (ceud_expl(r["explanatory"], CEUD_YEAR, "households", "single_detached") * 1e3) \
        * KWH_PER_GJ * hdd_uplift
    print(f"    -> reconciled stock implies {implied:,.0f} residential dwellings "
          f"vs {ott_hh:,} census households ({implied/ott_hh:.2f}x, target <=1.10x). "
          f"Per-UNIT unchanged: modelled detached mean "
          f"{det['annual_kwh'].mean():,.0f} kWh vs CEUD single_detached "
          f"{ceud_det_kwh:,.0f} kWh ({100*(det['annual_kwh'].mean()-ceud_det_kwh)/ceud_det_kwh:+.0f}%).")

    # --- (b) gas-heated share vs StatCan --------------------------------------
    res_fuel = res["heat_fuel"].value_counts(normalize=True)
    all_fuel = gdf["heat_fuel"].value_counts(normalize=True)
    print(f"\n(b) Residential fuel shares (target gas {STATCAN_GAS_SHARE:.0%}, "
          f"electric {STATCAN_ELEC_SHARE:.0%}):")
    for f in FUELS:
        print(f"      {f:9s} {res_fuel.get(f,0)*100:5.1f}%   (all classes "
              f"{all_fuel.get(f,0)*100:4.1f}%)")

    # --- (c) vs community energy inventory (emissions) ------------------------
    # modelled fuel INPUT for space heat, and its emissions
    inp = gdf["annual_kwh"] / gdf["heat_fuel"].map(AFUE)
    emis_t = (inp * gdf["heat_fuel"].map(EF_KG_PER_KWH)).sum() / 1000.0
    gas_input_twh = inp[gdf["heat_fuel"] == "gas"].sum() / 1e9
    inv = 2738852.0
    print(f"\n(c) Community inventory cross-check (Energy Evolution 2024 "
          f"'Buildings' = {inv:,.0f} tCO2e):")
    print(f"      Modelled SPACE-HEAT emissions {emis_t:,.0f} tCO2e "
          f"= {100*emis_t/inv:.0f}% of the buildings inventory (target <100%: "
          f"space heat is only ~55-70% of building energy -- water heat, "
          f"appliances, lighting, cooling, commercial process all excluded here).")
    print(f"      Modelled natural-gas INPUT for space heat: {gas_input_twh:.2f} "
          f"TWh/yr (space heat is documented as the dominant gas end use).")

    # --- (d) EWRB calibration cross-check -------------------------------------
    dd = intens["_diag"]
    print(f"\n(d) EWRB actuals cross-check (no per-building override possible -- "
          f"EWRB has no street address, source notes §2.1):")
    print(f"      EWRB Ottawa median total Site_EUI: commercial "
          f"{dd['ewrb']['commercial']:.2f} GJ/m2 (n={dd['ewrb']['commercial_n']}), "
          f"institutional {dd['ewrb']['institutional']:.2f} "
          f"(n={dd['ewrb']['institutional_n']}), industrial "
          f"{dd['ewrb']['industrial']:.2f} (n={dd['ewrb']['industrial_n']}).")
    print(f"      CEUD provincial commercial intensity is "
          f"{dd['ceud_com_gj_m2']/(dd['ewrb']['commercial']*dd['ceud_com_space_share']):.2f}x "
          f"the EWRB-calibrated space-heat value -> EWRB actuals preferred as the "
          f"commercial base (plan's 'EWRB calibrates the commercial table').")

    # --- (e) independent intensity cross-checks -------------------------------
    print(f"\n(e) Independent per-type intensity cross-checks (HDD-normalised):")
    _crosscheck_ceud(gdf, arch)
    _crosscheck_tandm(gdf, arch)


def _crosscheck_ceud(gdf, arch):
    """Modelled Ottawa detached/row per-m2 vs CEUD ON single_detached/attached,
    HDD-normalised (Ottawa 4407 / ON populated ~3900 = 1.13)."""
    r = json.load(open(CEUD_RES))
    up = 4407.0 / 3900.0
    for cls_name, bt in [("detached", "single_detached"), ("row", "single_attached")]:
        sub = gdf[gdf["class"] == cls_name]
        # compare on HEATED area (CEUD floor space is heated interior), matching
        # the house method's gross->heated conversion.
        model_kwh_m2 = sub["annual_kwh"].sum() / (sub["floor_area_m2"].sum()
                                                  * HEATED_FRACTION_HOUSE)
        sh = ceud_record(r["records"], CEUD_YEAR, bt, "space_heating")
        fs = ceud_expl(r["explanatory"], CEUD_YEAR, "floor_space", bt)
        ceud_kwh_m2 = (sh / fs) * KWH_PER_GJ * up
        print(f"      {cls_name:9s}: modelled {model_kwh_m2:5.1f} vs CEUD-ON "
              f"(HDD-adj) {ceud_kwh_m2:5.1f} kWh/m2  "
              f"({100*(model_kwh_m2-ceud_kwh_m2)/ceud_kwh_m2:+.0f}%)")


def _crosscheck_tandm(gdf, arch):
    """TaNDM Kelowna 2021 inventory (github.com/canmet-energy/tandm) single
    detached: gas per gas-metered unit -> implied delivered space heat, HDD-
    normalised Kelowna(~3000)->Ottawa(4407). Kelowna is milder, so expect Ottawa
    higher once normalised. Rough cross-check (gas includes water heating)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(TANDM, read_only=True, data_only=True)
        ws = wb["TaNDM IR2"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(h).strip() for h in rows[0]]
        idx = {h: i for i, h in enumerate(hdr)}
        gj = idx["Natural_Gas_GJ_yr"]; gm = idx["Gas_Meter_Count"]
        sub = idx["Sub-Category"]
        tot_gj = tot_m = 0.0
        for row in rows[1:]:
            if row[sub] and str(row[sub]).strip() == "Single Detached":
                if isinstance(row[gj], (int, float)) and isinstance(row[gm], (int, float)):
                    tot_gj += row[gj]; tot_m += row[gm]
        gj_per_gasunit = tot_gj / tot_m
        # implied delivered space heat: gas input x space-heat fraction x AFUE
        space_frac, afue = 0.78, 0.90
        kel_deliv_kwh = gj_per_gasunit * space_frac * afue * KWH_PER_GJ
        hdd = 4407.0 / 3000.0
        kel_norm = kel_deliv_kwh * hdd
        det = gdf[gdf["class"] == "detached"]
        model_kwh_unit = det["annual_kwh"].mean()   # ~1 dwelling per detached
        print(f"      TaNDM Kelowna detached: {gj_per_gasunit:.1f} GJ gas/gas-unit "
              f"-> {kel_deliv_kwh:,.0f} kWh delivered; HDD-norm to Ottawa "
              f"{kel_norm:,.0f} kWh vs modelled detached mean "
              f"{model_kwh_unit:,.0f} kWh "
              f"({100*(model_kwh_unit-kel_norm)/kel_norm:+.0f}%).")
        print(f"      (Kelowna gas incl. water heating & assumes 78% space "
              f"fraction; milder-climate normalisation is approximate.)")
    except Exception as e:
        print(f"      TaNDM cross-check skipped ({e}).")


def write_outputs(gdf):
    print("\nWriting outputs ...")
    meta = {
        "layer": "Ottawa building heat demand (screening estimate)",
        "phase": "HEATDEMAND_PLAN.md Phase 2 (build_building_demand.py)",
        "warning": ("SCREENING ESTIMATE, not a building audit. annual_kwh is "
                    "modelled space-heat delivered energy; vintage/type are "
                    "probabilistic per building (Phase 1); heat_fuel is a raked "
                    "probabilistic draw. Meaningful in aggregate (500 m cell / "
                    "feeder), not for a single address. CITY-WIDE SUMS: take over "
                    "in_ottawa_cd==True only (the bbox extends into surrounding "
                    "townships; the census is city-only). Stock reconciled in "
                    "Phase 2.5 (build_building_stock reconcile_stock, README "
                    "§3.10): per-DA implied dwellings capped to census; excess "
                    "-> 'accessory' class (unheated here)."),
        "columns": ("annual_kwh=annual space-heat delivered kWh/yr; "
                    "design_kw=design-day heat loss kW @ -22.8C; "
                    "ua_w_per_k=effective envelope UA; heat_fuel=primary fuel; "
                    "units_est=estimated dwelling units (MURB only); "
                    "demand_method/demand_confidence=provenance flags; "
                    "in_ottawa_cd=inside City of Ottawa census division; "
                    "assign_path=Phase 1 class provenance"),
        "methods": ("houses: Ottawa ERS archetypes scaled by floor area; "
                    "MURB: CEUD apt/attached ratio x Ottawa row archetype; "
                    "non-res: EWRB Ottawa actual Site_EUI x CEUD commercial "
                    "space-heat share; fuel: ERS per-FSA raked to StatCan "
                    "38-10-0286 with rural no-gas constraint."),
    }
    # parquet with metadata
    import pyarrow.parquet as pq
    # geopandas parquet keeps geo metadata; attach our screening note too
    gdf.to_parquet(BUILDINGS)
    # re-open to append custom metadata
    t = pq.read_table(BUILDINGS)
    md = dict(t.schema.metadata or {})
    md[b"heatdemand_phase2"] = json.dumps(meta).encode()
    t = t.replace_schema_metadata(md)
    pq.write_table(t, BUILDINGS)
    print(f"  {BUILDINGS}  ({len(gdf):,} rows, {len(gdf.columns)} cols)")

    # gpkg for QGIS spot-checks (drop heavy object cols it can't type well)
    try:
        gdf.to_file(BUILD_GPKG, driver="GPKG", layer="buildings_ottawa")
        print(f"  {BUILD_GPKG}")
    except Exception as e:
        print(f"  gpkg skipped ({e})")


if __name__ == "__main__":
    main()
